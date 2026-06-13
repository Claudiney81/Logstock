from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    make_response,
    current_app
)

from flask_login import login_required, current_user

from datetime import datetime
from types import SimpleNamespace

from sqlalchemy import func

import io
import os
import pandas as pd

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from app import db

from app.models import (
    Tecnico,
    TipoServico,
    SaldoTecnico,
    BaixaTecnica,
    BaixaTecnicaItem,
    Item,
    Empresa,
    OrdemServico
)

from app.utils.mailer import (
    send_baixa_recusa_email,
    send_baixa_aprovada_email
)

bp_baixa_desktop = Blueprint(
    "baixa_desktop",
    __name__,
    url_prefix="/baixa_desktop"
)


def buscar_saldo_tecnico(
    tecnico_id,
    item_id,
    tipo_servico_id,
    tipo_estoque,
    cliente_id=None,
    ordem_servico_id=None
):
    """
    Busca o saldo do técnico respeitando a origem do item.

    Regra:
    - Instalação, Manutenção e Reparo consomem saldo técnico de Instalação.
    - Empresa: saldo geral do técnico.
    - Cliente: saldo específico do cliente e da O.S.
    """

    # O serviço da baixa classifica o documento, mas o saldo físico é Instalação.
    if tipo_servico_id:
        tipo_servico_id = 1

    query = SaldoTecnico.query.filter(
        SaldoTecnico.tecnico_id == tecnico_id,
        SaldoTecnico.item_id == item_id,
        SaldoTecnico.tipo_servico_id == tipo_servico_id,
        SaldoTecnico.tipo_estoque == tipo_estoque,
        SaldoTecnico.quantidade > 0
    )

    if tipo_estoque == "cliente":
        query = query.filter(
            SaldoTecnico.cliente_id == cliente_id,
            SaldoTecnico.ordem_servico_id == ordem_servico_id
        )

    return query.order_by(SaldoTecnico.id.asc()).all()


def saldo_disponivel(
    tecnico_id,
    item_id,
    tipo_servico_id,
    tipo_estoque,
    cliente_id=None,
    ordem_servico_id=None
):
    saldos = buscar_saldo_tecnico(
        tecnico_id=tecnico_id,
        item_id=item_id,
        tipo_servico_id=tipo_servico_id,
        tipo_estoque=tipo_estoque,
        cliente_id=cliente_id,
        ordem_servico_id=ordem_servico_id
    )

    return sum(int(s.quantidade or 0) for s in saldos)


def valor_saldo_tecnico(
    tecnico_id,
    item_id,
    tipo_servico_id,
    tipo_estoque,
    cliente_id=None,
    ordem_servico_id=None
):
    saldos = buscar_saldo_tecnico(
        tecnico_id=tecnico_id,
        item_id=item_id,
        tipo_servico_id=tipo_servico_id,
        tipo_estoque=tipo_estoque,
        cliente_id=cliente_id,
        ordem_servico_id=ordem_servico_id
    )

    for saldo in saldos:
        if saldo.valor_unitario is not None:
            return float(saldo.valor_unitario or 0)

    item = Item.query.get(item_id)
    return float(item.valor or 0) if item else 0


def debitar_saldo_tecnico(
    tecnico_id,
    item_id,
    tipo_servico_id,
    tipo_estoque,
    cliente_id,
    ordem_servico_id,
    quantidade
):
    saldos = buscar_saldo_tecnico(
        tecnico_id=tecnico_id,
        item_id=item_id,
        tipo_servico_id=tipo_servico_id,
        tipo_estoque=tipo_estoque,
        cliente_id=cliente_id,
        ordem_servico_id=ordem_servico_id
    )

    disponivel = sum(int(s.quantidade or 0) for s in saldos)

    if disponivel < quantidade:
        return False

    restante = quantidade

    for saldo in saldos:
        atual = int(saldo.quantidade or 0)

        if atual <= 0:
            continue

        if atual >= restante:
            saldo.quantidade = atual - restante
            restante = 0
            break

        saldo.quantidade = 0
        restante -= atual

    return restante == 0


@bp_baixa_desktop.route("/api/ordens-servico")
@login_required
def api_ordens_servico():
    cliente_id = request.args.get("cliente_id", type=int)

    if not cliente_id:
        return {"ordens": []}

    # A O.S é para REGISTRO da baixa.
    # Por isso NÃO pode depender de saldo técnico.
    # Lista todas as O.S cadastradas para o cliente selecionado.
    ordens = (
        OrdemServico.query
        .filter(OrdemServico.cliente_id == cliente_id)
        .order_by(OrdemServico.numero_os.asc())
        .all()
    )

    return {
        "ordens": [
            {
                "id": os.id,
                "numero_os": os.numero_os,
                "endereco": os.endereco or ""
            }
            for os in ordens
        ]
    }

@bp_baixa_desktop.route("/api/itens-saldo")
@login_required
def api_itens_saldo():
    tecnico_id = request.args.get("tecnico_id", type=int)
    tipo_servico_id = request.args.get("tipo_servico_id", type=int)
    tipo_estoque = (request.args.get("tipo_estoque") or "").strip()
    cliente_id = request.args.get("cliente_id", type=int)
    ordem_servico_id = request.args.get("ordem_servico_id", type=int)

    if not tecnico_id or not tipo_servico_id or tipo_estoque not in ["empresa", "cliente"]:
        return {"itens": []}

    # O serviço selecionado classifica a baixa; o saldo físico vem da Instalação.
    tipo_servico_consulta = 1

    query = (
        db.session.query(
            SaldoTecnico.item_id,
            func.sum(SaldoTecnico.quantidade).label("quantidade")
        )
        .filter(
            SaldoTecnico.tecnico_id == tecnico_id,
            SaldoTecnico.tipo_servico_id == tipo_servico_consulta,
            SaldoTecnico.tipo_estoque == tipo_estoque,
            SaldoTecnico.quantidade > 0
        )
    )

    if tipo_estoque == "cliente":
        if not cliente_id or not ordem_servico_id:
            return {"itens": []}

        query = query.filter(
            SaldoTecnico.cliente_id == cliente_id,
            SaldoTecnico.ordem_servico_id == ordem_servico_id
        )

    elif tipo_estoque == "empresa":
        query = query.filter(
            SaldoTecnico.cliente_id.is_(None),
            SaldoTecnico.ordem_servico_id.is_(None)
        )

    rows = (
        query
        .group_by(SaldoTecnico.item_id)
        .order_by(SaldoTecnico.item_id.asc())
        .all()
    )

    itens = []

    for item_id, quantidade in rows:
        item = Item.query.get(item_id)

        if item and int(quantidade or 0) > 0:
            itens.append({
                "item_id": item.id,
                "codigo": item.codigo,
                "descricao": item.descricao,
                "unidade": item.unidade,
                "quantidade": int(quantidade or 0)
            })

    return {"itens": itens}
@bp_baixa_desktop.route("/nova", methods=["GET", "POST"])
@login_required
def nova_baixa():
    tecnicos = Tecnico.query.order_by(Tecnico.nome).all()
    tipos_servico = TipoServico.query.order_by(TipoServico.nome).all()

    clientes = (
        Empresa.query
        .filter(db.func.lower(Empresa.tipo_empresa) == "cliente")
        .order_by(Empresa.razao_social)
        .all()
    )

    if request.method == "POST":
        baixa_id_corrigir = request.form.get("baixa_id_corrigir", type=int)

        tecnico_id = request.form.get("tecnico_id", type=int)
        tipo_servico_id = request.form.get("tipo_servico_id", type=int)
        cliente_id = request.form.get("cliente_id", type=int)
        ordem_servico_id = request.form.get("ordem_servico_id", type=int)
        observacao = request.form.get("observacao", "").strip()

        if not tecnico_id or not tipo_servico_id or not cliente_id or not ordem_servico_id:
            flash("Preencha Técnico, Tipo de Serviço, Cliente e O.S.", "warning")
            return redirect(url_for("baixa_desktop.nova_baixa"))

        cliente = Empresa.query.get(cliente_id)
        ordem_servico = OrdemServico.query.get(ordem_servico_id)

        item_ids = request.form.getlist("resumo_item_id[]")
        quantidades = request.form.getlist("resumo_quantidade[]")
        tipos_estoque = request.form.getlist("resumo_tipo_estoque[]")

        itens_compilados = []

        for item_id, qtd, tipo_estoque_item in zip(item_ids, quantidades, tipos_estoque):
            try:
                item_id = int(item_id)
                qtd = int(qtd or 0)
            except Exception:
                continue

            tipo_estoque_item = (tipo_estoque_item or "").strip()

            if qtd <= 0 or tipo_estoque_item not in ["empresa", "cliente"]:
                continue

            cliente_ref = cliente_id if tipo_estoque_item == "cliente" else None
            os_ref = ordem_servico_id if tipo_estoque_item == "cliente" else None

            saldo_total = saldo_disponivel(
                tecnico_id=tecnico_id,
                item_id=item_id,
                tipo_servico_id=tipo_servico_id,
                tipo_estoque=tipo_estoque_item,
                cliente_id=cliente_ref,
                ordem_servico_id=os_ref
            )

            if qtd > saldo_total:
                item = Item.query.get(item_id)
                flash(
                    f"Quantidade maior que o saldo disponível para {item.descricao if item else item_id}.",
                    "danger"
                )
                return redirect(url_for("baixa_desktop.nova_baixa"))

            itens_compilados.append({
                "item_id": item_id,
                "quantidade": qtd,
                "tipo_estoque": tipo_estoque_item,
                "cliente_estoque_id": cliente_ref
            })

        if not itens_compilados:
            flash("Adicione pelo menos um item ao resumo da baixa.", "warning")
            return redirect(url_for("baixa_desktop.nova_baixa"))

        responsavel = (
            getattr(current_user, "nome", None)
            or getattr(current_user, "username", None)
            or getattr(current_user, "email", None)
            or "Usuário Logado"
        )

        if baixa_id_corrigir:
            baixa = BaixaTecnica.query.filter(
                BaixaTecnica.id == baixa_id_corrigir,
                BaixaTecnica.status == "pendente_ajuste",
                BaixaTecnica.origem_mobile == False
            ).first()

            if not baixa:
                flash("Baixa devolvida não encontrada ou já finalizada.", "danger")
                return redirect(url_for("baixa_desktop.nova_baixa"))

            # Remove somente os itens que ainda estavam pendentes/devolvidos para ajuste.
            # Itens já confirmados em aprovação parcial não são apagados.
            for item_antigo in list(baixa.itens):
                if item_antigo.status in ["pendente", "pendente_ajuste", "recusado"]:
                    db.session.delete(item_antigo)

            baixa.tecnico_id = tecnico_id
            baixa.tipo_servico_id = tipo_servico_id
            baixa.cliente_id = cliente_id
            baixa.ordem_servico_id = ordem_servico_id
            baixa.endereco = ordem_servico.endereco if ordem_servico else None
            baixa.os_cliente = (
                f"{cliente.razao_social} - {ordem_servico.numero_os}"
                if cliente and ordem_servico
                else None
            )
            baixa.responsavel = responsavel
            baixa.observacao = observacao
            baixa.status = "pendente"
            baixa.motivo_recusa = None
            baixa.visualizado_tecnico = True
            baixa.data_hora = datetime.now()

            mensagem_sucesso = "Baixa corrigida e reenviada para aprovação do engenheiro/supervisor."

        else:
            baixa = BaixaTecnica(
                tecnico_id=tecnico_id,
                tipo_servico_id=tipo_servico_id,
                cliente_id=cliente_id,
                ordem_servico_id=ordem_servico_id,
                endereco=ordem_servico.endereco if ordem_servico else None,
                os_cliente=(
                    f"{cliente.razao_social} - {ordem_servico.numero_os}"
                    if cliente and ordem_servico
                    else None
                ),
                responsavel=responsavel,
                observacao=observacao,
                status="pendente",
                origem_mobile=False,
                visualizado_tecnico=True,
                data_hora=datetime.now()
            )

            db.session.add(baixa)
            db.session.flush()

            mensagem_sucesso = "Baixa registrada. Aguarde aprovação do engenheiro/supervisor."

        for linha in itens_compilados:
            valor_unitario = valor_saldo_tecnico(
                tecnico_id=tecnico_id,
                item_id=linha["item_id"],
                tipo_servico_id=tipo_servico_id,
                tipo_estoque=linha["tipo_estoque"],
                cliente_id=linha["cliente_estoque_id"],
                ordem_servico_id=(
                    ordem_servico_id
                    if linha["tipo_estoque"] == "cliente"
                    else None
                )
            )
            valor_total = linha["quantidade"] * valor_unitario

            db.session.add(BaixaTecnicaItem(
                baixa_tecnica_id=baixa.id,
                item_id=linha["item_id"],
                tipo_estoque=linha["tipo_estoque"],
                cliente_estoque_id=linha["cliente_estoque_id"],
                quantidade=linha["quantidade"],
                quantidade_aprovada=0,
                valor_unitario=valor_unitario,
                valor_total=valor_total,
                status="pendente"
            ))

        db.session.commit()

        flash(mensagem_sucesso, "success")
        return redirect(url_for("baixa_desktop.baixas_pendentes"))

    baixas_recusadas = (
        BaixaTecnica.query
        .filter(
            BaixaTecnica.status == "pendente_ajuste",
            BaixaTecnica.origem_mobile == False
        )
        .order_by(BaixaTecnica.data_hora.desc())
        .all()
    )

    # Para a correção na tela desktop:
    # mostra somente os itens devolvidos daquela baixa, mas exibindo o saldo real atual.
    for baixa_recusada in baixas_recusadas:
        for item_baixa in baixa_recusada.itens:
            if item_baixa.status not in ["pendente", "pendente_ajuste", "recusado"]:
                item_baixa.saldo_real_corrigir = 0
                continue

            cliente_ref = (
                item_baixa.cliente_estoque_id
                if item_baixa.tipo_estoque == "cliente"
                else None
            )

            ordem_ref = (
                baixa_recusada.ordem_servico_id
                if item_baixa.tipo_estoque == "cliente"
                else None
            )

            item_baixa.saldo_real_corrigir = saldo_disponivel(
                tecnico_id=baixa_recusada.tecnico_id,
                item_id=item_baixa.item_id,
                tipo_servico_id=baixa_recusada.tipo_servico_id,
                tipo_estoque=item_baixa.tipo_estoque or "empresa",
                cliente_id=cliente_ref,
                ordem_servico_id=ordem_ref
            )

    return render_template(
        "baixa_desktop/nova_baixa.html",
        tecnicos=tecnicos,
        tipos_servico=tipos_servico,
        clientes=clientes,
        baixas_recusadas=baixas_recusadas
    )


@bp_baixa_desktop.route("/detalhe/<int:baixa_id>", methods=["GET", "POST"])
@login_required
def detalhe_baixa(baixa_id):
    baixa = BaixaTecnica.query.get_or_404(baixa_id)

    itens = []

    for item_baixa in baixa.itens:
        cliente_ref = (
        item_baixa.cliente_estoque_id
        if item_baixa.tipo_estoque == "cliente"
        else None
    )

        saldo = saldo_disponivel(
            tecnico_id=baixa.tecnico_id,
            item_id=item_baixa.item_id,
            tipo_servico_id=baixa.tipo_servico_id,
            tipo_estoque=item_baixa.tipo_estoque or "empresa",
            cliente_id=cliente_ref,
            ordem_servico_id=(
                baixa.ordem_servico_id
                if item_baixa.tipo_estoque == "cliente"
                else None
            )
        )

        itens.append((item_baixa, saldo))

    if request.method == "POST":

        if baixa.status in ["confirmado", "recusado"]:
            flash("Esta baixa já foi finalizada.", "info")
            return redirect(url_for("baixa_desktop.baixas_pendentes"))

        itens_ids = [
            int(x)
            for x in request.form.getlist("itens")
            if str(x).isdigit()
        ]

        qtd_digitada = {
            int(k.split("_", 1)[1]): int(v) if str(v).isdigit() else 0
            for k, v in request.form.items()
            if k.startswith("qtd_")
        }

        # ==================================================
        # APROVAR BAIXA
        # ==================================================
        if "aprovar" in request.form:

            if not itens_ids:
                flash("Selecione pelo menos um item para aprovar.", "warning")
                return redirect(url_for("baixa_desktop.detalhe_baixa", baixa_id=baixa.id))

            selecionados = (
                BaixaTecnicaItem.query
                .filter(
                    BaixaTecnicaItem.baixa_tecnica_id == baixa.id,
                    BaixaTecnicaItem.id.in_(itens_ids),
                    BaixaTecnicaItem.status == "pendente"
                )
                .all()
            )

            erros = []
            aprovacoes = []

            for item_baixa in selecionados:
                solicitado = int(item_baixa.quantidade or 0)
                qtd_aprovar = int(qtd_digitada.get(item_baixa.id, solicitado) or 0)

                cliente_ref = (
                    item_baixa.cliente_estoque_id
                    if item_baixa.tipo_estoque == "cliente"
                    else None
                )

                disponivel = saldo_disponivel(
                    tecnico_id=baixa.tecnico_id,
                    item_id=item_baixa.item_id,
                    tipo_servico_id=baixa.tipo_servico_id,
                    tipo_estoque=item_baixa.tipo_estoque or "empresa",
                    cliente_id=cliente_ref,
                    ordem_servico_id=(
                        baixa.ordem_servico_id
                        if item_baixa.tipo_estoque == "cliente"
                        else None
                    )
                )

                maximo = min(solicitado, disponivel)

                if qtd_aprovar <= 0 or qtd_aprovar > maximo:
                    codigo = getattr(item_baixa.item, "codigo", item_baixa.item_id)
                    erros.append(f"Item {codigo}: informe entre 1 e {maximo}.")
                    continue

                aprovacoes.append((item_baixa, qtd_aprovar, cliente_ref))

            if erros:
                flash("Não foi possível aprovar: " + " ".join(erros), "danger")
                return redirect(url_for("baixa_desktop.detalhe_baixa", baixa_id=baixa.id))

            for item_baixa, qtd_aprovar, cliente_ref in aprovacoes:

                ok = debitar_saldo_tecnico(
                    tecnico_id=baixa.tecnico_id,
                    item_id=item_baixa.item_id,
                    tipo_servico_id=baixa.tipo_servico_id,
                    tipo_estoque=item_baixa.tipo_estoque or "empresa",
                    cliente_id=cliente_ref,
                    ordem_servico_id=(
                        baixa.ordem_servico_id
                        if item_baixa.tipo_estoque == "cliente"
                        else None
                    ),
                    quantidade=qtd_aprovar
                )

                if not ok:
                    db.session.rollback()
                    codigo = getattr(item_baixa.item, "codigo", item_baixa.item_id)
                    flash(f"Erro ao debitar item {codigo}.", "danger")
                    return redirect(url_for("baixa_desktop.detalhe_baixa", baixa_id=baixa.id))

                item_baixa.quantidade_aprovada += qtd_aprovar
                item_baixa.quantidade -= qtd_aprovar

                if item_baixa.quantidade <= 0:
                    item_baixa.quantidade = 0
                    item_baixa.status = "confirmado"
                else:
                    item_baixa.status = "pendente"

            pendentes = (
                BaixaTecnicaItem.query
                .filter_by(
                    baixa_tecnica_id=baixa.id,
                    status="pendente"
                )
                .count()
            )

            if pendentes == 0:
                baixa.status = "confirmado"
                baixa.visualizado_tecnico = True
            else:
                baixa.status = "pendente"

            if baixa.cliente_id:
                cliente_os = Empresa.query.get(baixa.cliente_id)

                if cliente_os:
                    baixas_abertas = (
                        BaixaTecnica.query
                        .filter(
                            BaixaTecnica.cliente_id == baixa.cliente_id,
                            BaixaTecnica.tecnico_id == baixa.tecnico_id,
                            BaixaTecnica.status.in_(["pendente", "pendente_ajuste"])
                        )
                        .count()
                    )

                    cliente_os.status_os = (
                        "finalizada"
                        if baixas_abertas == 0
                        else "em_andamento"
                    )

            db.session.commit()

            flash("Baixa aprovada com sucesso.", "success")
            return redirect(url_for("baixa_desktop.baixas_pendentes"))

        # ==================================================
        # RECUSAR / DEVOLVER BAIXA
        # ==================================================
        if "recusar" in request.form:

            motivo = (request.form.get("motivo") or "").strip()

            if not motivo:
                flash("Informe o motivo da devolução para ajuste.", "warning")
                return redirect(url_for("baixa_desktop.detalhe_baixa", baixa_id=baixa.id))

            itens_devolvidos = (
                BaixaTecnicaItem.query
                .filter(
                    BaixaTecnicaItem.baixa_tecnica_id == baixa.id,
                    BaixaTecnicaItem.status == "pendente"
                )
                .all()
            )

            if not itens_devolvidos:
                flash("Nenhum item pendente encontrado para devolução.", "warning")
                return redirect(url_for("baixa_desktop.detalhe_baixa", baixa_id=baixa.id))

            baixa.status = "pendente_ajuste"
            baixa.motivo_recusa = motivo
            baixa.visualizado_tecnico = False

            for item_baixa in itens_devolvidos:
                item_baixa.status = "pendente_ajuste"
                item_baixa.motivo_recusa = motivo

            if baixa.cliente_id:
                cliente_os = Empresa.query.get(baixa.cliente_id)

                if cliente_os:
                    cliente_os.status_os = "em_andamento"

            db.session.commit()

            flash("Baixa devolvida para ajuste do técnico.", "warning")
            return redirect(url_for("baixa_desktop.baixas_pendentes"))

    return render_template(
        "baixa_desktop/detalhe_baixa.html",
        baixa=baixa,
        itens=itens
    )
    
@bp_baixa_desktop.route("/pendentes")
@login_required
def baixas_pendentes():

    baixas = (
        BaixaTecnica.query
        .filter(
            BaixaTecnica.status.in_(["pendente", "pendente_ajuste"])
        )
        .order_by(BaixaTecnica.data_hora.desc())
        .all()
    )

    return render_template(
        "baixa_desktop/baixas_pendentes.html",
        baixas=baixas
    )
    
@bp_baixa_desktop.route("/historico")
@login_required
def historico_baixas():
    tecnicos = Tecnico.query.order_by(Tecnico.nome).all()
    clientes = Empresa.query.order_by(Empresa.razao_social).all()
    tipos_servico = TipoServico.query.order_by(TipoServico.nome).all()

    tecnico_id = request.args.get("tecnico_id", type=int)
    cliente_id = request.args.get("cliente_id", type=int)
    tipo_servico_id = request.args.get("tipo_servico_id", type=int)
    tipo_estoque = request.args.get("tipo_estoque", "")
    status = request.args.get("status", "")
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    query = BaixaTecnica.query

    if tecnico_id:
        query = query.filter(BaixaTecnica.tecnico_id == tecnico_id)

    if cliente_id:
        query = query.filter(BaixaTecnica.cliente_id == cliente_id)

    if tipo_servico_id:
        query = query.filter(BaixaTecnica.tipo_servico_id == tipo_servico_id)

    if status:
        query = query.filter(BaixaTecnica.status == status)

    if tipo_estoque:
        query = query.filter(
            BaixaTecnica.itens.any(
                BaixaTecnicaItem.tipo_estoque == tipo_estoque
            )
        )

    if data_inicio:
        inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
        query = query.filter(BaixaTecnica.data_hora >= inicio)

    if data_fim:
        fim = datetime.strptime(data_fim, "%Y-%m-%d")
        fim = fim.replace(hour=23, minute=59, second=59)
        query = query.filter(BaixaTecnica.data_hora <= fim)

    historico = query.order_by(BaixaTecnica.data_hora.desc()).all()

    return render_template(
        "baixa_desktop/historico_baixas.html",
        historico=historico,
        tecnicos=tecnicos,
        clientes=clientes,
        tipos_servico=tipos_servico,
        tecnico_id=tecnico_id,
        cliente_id=cliente_id,
        tipo_servico_id=tipo_servico_id,
        tipo_estoque=tipo_estoque,
        status=status,
        data_inicio=data_inicio,
        data_fim=data_fim
    )


@bp_baixa_desktop.route("/realizadas")
@login_required
def baixas_realizadas():
    tecnicos = Tecnico.query.order_by(Tecnico.nome).all()
    clientes = Empresa.query.filter(
        db.func.lower(Empresa.tipo_empresa) == "cliente"
    ).order_by(Empresa.razao_social).all()
    tipos_servico = TipoServico.query.order_by(TipoServico.nome).all()

    tecnico_id = request.args.get("tecnico_id", type=int)
    cliente_id = request.args.get("cliente_id", type=int)
    tipo_servico_id = request.args.get("tipo_servico_id", type=int)
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    query = BaixaTecnica.query.filter_by(status="confirmado")

    if tecnico_id:
        query = query.filter(BaixaTecnica.tecnico_id == tecnico_id)

    if cliente_id:
        query = query.filter(BaixaTecnica.cliente_id == cliente_id)

    if tipo_servico_id:
        query = query.filter(BaixaTecnica.tipo_servico_id == tipo_servico_id)

    if data_inicio:
        query = query.filter(
            BaixaTecnica.data_hora >= datetime.strptime(data_inicio, "%Y-%m-%d")
        )

    if data_fim:
        fim = datetime.strptime(data_fim, "%Y-%m-%d")
        fim = fim.replace(hour=23, minute=59, second=59)
        query = query.filter(BaixaTecnica.data_hora <= fim)

    baixas = query.order_by(BaixaTecnica.data_hora.desc()).all()

    
    return render_template(
        "baixa_desktop/baixas_realizadas.html",
        baixas=baixas,
        tecnicos=tecnicos,
        clientes=clientes,
        tipos_servico=tipos_servico,
        tecnico_id=tecnico_id,
        cliente_id=cliente_id,
        tipo_servico_id=tipo_servico_id,
        data_inicio=data_inicio,
        data_fim=data_fim
    )       

@bp_baixa_desktop.route("/detalhe/<int:baixa_id>/excel")
@login_required
def exportar_baixa_excel(baixa_id):

    baixa = BaixaTecnica.query.get_or_404(baixa_id)

    dados = []

    total_geral = 0
    total_empresa = 0
    total_cliente = 0

    for item in baixa.itens:

        valor_unit = float(item.valor_unitario or 0)
        valor_total = float(item.valor_total or 0)

        total_geral += valor_total

        if (item.tipo_estoque or "").lower() == "empresa":
            total_empresa += valor_total
        elif (item.tipo_estoque or "").lower() == "cliente":
            total_cliente += valor_total

        dados.append({
            "Código": item.item.codigo if item.item else "",
            "Descrição": item.item.descricao if item.item else "",
            "Unidade": item.item.unidade if item.item else "",
            "Quantidade": item.quantidade_aprovada or item.quantidade,
            "Valor Unitário": valor_unit,
            "Valor Total": valor_total,
            "Tipo Estoque": (item.tipo_estoque or "").title()
        })

    df = pd.DataFrame(dados)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        sheet_name = "Baixa Técnica"

        df.to_excel(
            writer,
            index=False,
            startrow=10,
            sheet_name=sheet_name
        )

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A11"

        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        worksheet.page_margins.left = 0.3
        worksheet.page_margins.right = 0.3
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5

        azul = PatternFill(start_color="002B55", end_color="002B55", fill_type="solid")
        azul_claro = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
        cinza_claro = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        branco = Font(color="FFFFFF", bold=True)
        titulo_font = Font(bold=True, size=15, color="FFFFFF")
        label_font = Font(bold=True, color="002B55")
        texto_font = Font(color="111827")

        borda_fina = Side(style="thin", color="D1D5DB")
        borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

        worksheet.merge_cells("A1:G1")
        worksheet["A1"] = "RELATÓRIO DE MATERIAL APLICADO"
        worksheet["A1"].font = titulo_font
        worksheet["A1"].fill = azul
        worksheet["A1"].alignment = center
        worksheet.row_dimensions[1].height = 26

        cliente_nome = baixa.cliente.razao_social if baixa.cliente else "-"

        numero_os = "-"
        if baixa.ordem_servico:
            numero_os = baixa.ordem_servico.numero_os or "-"
        elif baixa.os_cliente:
            numero_os = baixa.os_cliente

        dados_cabecalho = {
            "A3": "Técnico:",
            "B3": baixa.tecnico.nome if baixa.tecnico else "-",
            "D3": "Data:",
            "E3": baixa.data_hora.strftime("%d/%m/%Y %H:%M") if baixa.data_hora else "-",
            "A4": "Cliente:",
            "B4": cliente_nome,
            "D4": "O.S:",
            "E4": numero_os,
            "A5": "Tipo Serviço:",
            "B5": baixa.tipo_servico.nome if baixa.tipo_servico else "-",
            "D5": "Responsável:",
            "E5": baixa.responsavel or "-",
            "A6": "Endereço:",
            "B6": baixa.endereco or "-",
            "A7": "Observação:",
            "B7": baixa.observacao or "-"
        }

        for celula, valor in dados_cabecalho.items():
            worksheet[celula] = valor

        worksheet.merge_cells("B6:G6")
        worksheet.merge_cells("B7:G7")

        for row in range(3, 8):
            for col in range(1, 8):
                cell = worksheet.cell(row=row, column=col)
                cell.border = borda
                cell.alignment = wrap if col in [2, 5] else left
                cell.font = texto_font
                cell.fill = cinza_claro

        for celula in ["A3", "D3", "A4", "D4", "A5", "D5", "A6", "A7"]:
            worksheet[celula].font = label_font
            worksheet[celula].fill = azul_claro

        worksheet.row_dimensions[6].height = 24
        worksheet.row_dimensions[7].height = 24

        header_row = 11

        for cell in worksheet[header_row]:
            cell.fill = azul
            cell.font = branco
            cell.alignment = center
            cell.border = borda

        worksheet.row_dimensions[header_row].height = 22

        primeira_linha = 12
        ultima_linha = 11 + len(df)

        for row in worksheet.iter_rows(
            min_row=primeira_linha,
            max_row=ultima_linha,
            min_col=1,
            max_col=7
        ):
            for cell in row:
                cell.border = borda
                cell.alignment = center
                cell.font = texto_font
                cell.fill = PatternFill(
                    start_color="FFFFFF" if cell.row % 2 == 0 else "F8FAFC",
                    end_color="FFFFFF" if cell.row % 2 == 0 else "F8FAFC",
                    fill_type="solid"
                )

            row[1].alignment = wrap
            row[4].number_format = 'R$ #,##0.00'
            row[5].number_format = 'R$ #,##0.00'
            row[4].alignment = right
            row[5].alignment = right

        larguras = {
            1: 15,
            2: 42,
            3: 10,
            4: 11,
            5: 14,
            6: 14,
            7: 14
        }

        for col, largura in larguras.items():
            worksheet.column_dimensions[get_column_letter(col)].width = largura

        linha_total_empresa = ultima_linha + 3
        linha_total_cliente = ultima_linha + 4
        linha_total_geral = ultima_linha + 6

        totais = [
            (linha_total_empresa, "TOTAL EMPRESA", total_empresa),
            (linha_total_cliente, "TOTAL CLIENTE", total_cliente),
            (linha_total_geral, "TOTAL GERAL DA BAIXA", total_geral),
        ]

        for linha, titulo, valor in totais:
            worksheet.merge_cells(
                start_row=linha,
                start_column=1,
                end_row=linha,
                end_column=4
            )

            worksheet.cell(row=linha, column=1, value=titulo)

            worksheet.merge_cells(
                start_row=linha,
                start_column=5,
                end_row=linha,
                end_column=7
            )

            worksheet.cell(row=linha, column=5, value=valor)

            for col in range(1, 8):
                cell = worksheet.cell(row=linha, column=col)
                cell.fill = azul
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.border = borda
                cell.alignment = center

            worksheet.cell(row=linha, column=5).number_format = 'R$ #,##0.00'

        linha_rodape = linha_total_geral + 3

        worksheet.merge_cells(
            start_row=linha_rodape,
            start_column=1,
            end_row=linha_rodape,
            end_column=7
        )

        worksheet.cell(
            row=linha_rodape,
            column=1,
            value="Documento gerado automaticamente pelo LogiStock."
        )

        worksheet.cell(row=linha_rodape, column=1).font = Font(
            italic=True,
            color="6B7280",
            size=9
        )

        worksheet.cell(row=linha_rodape, column=1).alignment = center

    output.seek(0)

    nome_arquivo = f"baixa_tecnica_{baixa.id}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@bp_baixa_desktop.route("/detalhe/<int:baixa_id>/pdf")
@login_required
def exportar_baixa_pdf(baixa_id):

    baixa = BaixaTecnica.query.get_or_404(baixa_id)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=25,
        leftMargin=25,
        topMargin=22,
        bottomMargin=22
    )

    elementos = []

    styles = getSampleStyleSheet()

    azul = colors.HexColor("#002b55")
    cinza_claro = colors.HexColor("#f3f6f9")
    cinza_borda = colors.HexColor("#c9c9c9")
    texto = colors.HexColor("#1f2937")

    def moeda(valor):
        return f"R$ {float(valor or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    # ==================================================
    # LOGO + CABEÇALHO
    # ==================================================

    logo_path = os.path.join(
        current_app.root_path,
        "static",
        "img",
        "start_logo.png"
    )

    logo = ""

    if os.path.exists(logo_path):
        logo = Image(logo_path, width=70, height=45)

    titulo = Paragraph(
        """
        <para align="right">
            <font size="18" color="#002b55">
                <b>RELATÓRIO DE MATERIAL APLICADO</b>
            </font><br/>
            <font size="9" color="#374151">
                Documento gerencial para análise de materiais aplicados
            </font>
        </para>
        """,
        styles["Normal"]
    )

    cabecalho = Table(
        [[logo, titulo]],
        colWidths=[120, 420]
    )

    cabecalho.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 2, azul),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))

    elementos.append(cabecalho)
    elementos.append(Spacer(1, 18))

    # ==================================================
    # CLIENTE / OS
    # ==================================================

    # ==================================================
# CLIENTE / OS
# ==================================================

    cliente_nome = "-"
    numero_os = "-"

    if baixa.cliente:
        cliente_nome = baixa.cliente.razao_social or "-"

    ordem = None

    if baixa.ordem_servico_id:
        ordem = OrdemServico.query.get(baixa.ordem_servico_id)

    if ordem:
        numero_os = ordem.numero_os or "-"
    elif baixa.os_cliente:
        numero_os = baixa.os_cliente

    if cliente_nome != "-":
        cliente_os = f"{cliente_nome} - {numero_os}"
    else:
        cliente_os = numero_os
# ==================================================
# DADOS DA APLICAÇÃO
# ==================================================

    status = baixa.status.replace("_", " ").capitalize() if baixa.status else "-"

    tipos_estoque = set()

    for item in baixa.itens:
        if item.tipo_estoque:
            tipos_estoque.add(item.tipo_estoque)

    if "empresa" in tipos_estoque and "cliente" in tipos_estoque:
        tipo_estoque_pdf = "Empresa / Cliente"
    elif "empresa" in tipos_estoque:
        tipo_estoque_pdf = "Empresa"
    elif "cliente" in tipos_estoque:
        tipo_estoque_pdf = "Cliente"
    else:
        tipo_estoque_pdf = "-"

    dados_cabecalho = [
    [
        "Técnico:",
        baixa.tecnico.nome if baixa.tecnico else "-",
        "Responsável:",
        baixa.responsavel or "-"
    ],

    [
        "Cliente:",
        cliente_nome,
        "O.S:",
        numero_os
    ],

    [
        "Tipo Serviço:",
        baixa.tipo_servico.nome if baixa.tipo_servico else "-",
        "Status:",
        status
    ],

    [
        "Tipo de Estoque:",
        tipo_estoque_pdf,
        "Endereço:",
        baixa.endereco or "-"
    ],

    [
        "Observação:",
        baixa.observacao or "-",
        "",
        ""
    ],
]

    titulo_dados = Table(
        [["DADOS DA APLICAÇÃO"]],
        colWidths=[540]
    )

    titulo_dados.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), azul),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))

    elementos.append(titulo_dados)

    tabela_cab = Table(
        dados_cabecalho,
        colWidths=[95, 150, 95, 200]
    )

    tabela_cab.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, cinza_borda),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("TEXTCOLOR", (0, 0), (-1, -1), texto),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), azul),
        ("TEXTCOLOR", (2, 0), (2, -1), azul),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabela_cab)
    elementos.append(Spacer(1, 18))

    # ==================================================
    # ITENS
    # ==================================================

    dados_itens = [[
        "Código",
        "Descrição",
        "Un.",
        "Qtd.",
        "Valor Unit.",
        "Valor Total",
        "Origem"
    ]]

    total_geral = 0
    total_empresa = 0
    total_cliente = 0

    for item in baixa.itens:

        valor_unit = float(item.valor_unitario or 0)
        valor_total = float(item.valor_total or 0)

        total_geral += valor_total

        if (item.tipo_estoque or "").lower() == "empresa":
            total_empresa += valor_total

        elif (item.tipo_estoque or "").lower() == "cliente":
            total_cliente += valor_total

        origem = "Cliente" if item.tipo_estoque == "cliente" else "Empresa"

        dados_itens.append([
            item.item.codigo if item.item else "-",
            Paragraph(item.item.descricao if item.item else "-", styles["Normal"]),
            item.item.unidade if item.item else "-",
            str(item.quantidade_aprovada or item.quantidade or 0),
            moeda(valor_unit),
            moeda(valor_total),
            origem
        ])

    tabela_itens = Table(
        dados_itens,
        colWidths=[62, 205, 38, 38, 70, 75, 52],
        repeatRows=1
    )

    tabela_itens.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), azul),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),

        ("GRID", (0, 0), (-1, -1), 0.5, cinza_borda),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (3, -1), "CENTER"),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("ALIGN", (6, 1), (6, -1), "CENTER"),

        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, cinza_claro]),
    ]))

    elementos.append(tabela_itens)
    elementos.append(Spacer(1, 18))

# ==================================================
# TOTAIS
# ==================================================

    dados_totais = [
        [
            Paragraph(
                '<para align="right"><b>TOTAL EMPRESA:</b></para>',
                styles["Normal"]
            ),
            Paragraph(
                f'<para align="center"><font color="white"><b>{moeda(total_empresa)}</b></font></para>',
                styles["Normal"]
            )
        ],
        [
            Paragraph(
                '<para align="right"><b>TOTAL CLIENTE:</b></para>',
                styles["Normal"]
            ),
            Paragraph(
                f'<para align="center"><font color="white"><b>{moeda(total_cliente)}</b></font></para>',
                styles["Normal"]
            )
        ],
        [
            Paragraph(
                '<para align="right"><b>TOTAL GERAL DA BAIXA:</b></para>',
                styles["Normal"]
            ),
            Paragraph(
                f'<para align="center"><font color="white" size="14"><b>{moeda(total_geral)}</b></font></para>',
                styles["Normal"]
            )
        ]
    ]

    tabela_total = Table(
        dados_totais,
        colWidths=[360, 180]
    )

    tabela_total.setStyle(TableStyle([
        ("BACKGROUND", (1, 0), (1, -1), azul),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.6, cinza_borda),
        ("GRID", (0, 0), (-1, -1), 0.4, cinza_borda),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (0, -1), 12),
    ]))

    elementos.append(tabela_total)

    elementos.append(Spacer(1, 24))

    assinatura_titulo = Paragraph(
        '<para align="center"><b>Assinatura do Responsável / Operador</b></para>',
        styles["Normal"]
    )
    assinatura_fisica = Paragraph(
        '<para align="center"><b>Assinatura física</b></para>',
        styles["Normal"]
    )
    assinatura_nome = baixa.responsavel or "Responsável / Operador"
    tabela_assinatura = Table([[assinatura_nome]], colWidths=[360])
    tabela_assinatura.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.HexColor("#374151")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(assinatura_titulo)
    elementos.append(Spacer(1, 12))
    elementos.append(assinatura_fisica)
    elementos.append(Spacer(1, 12))
    elementos.append(tabela_assinatura)

    # ==================================================
    # RODAPÉ PADRÃO LOGISTOCK
    # ==================================================

    elementos.append(Spacer(1, 28))

    rodape = Paragraph(
        """
        <para align="center">
            <font size="8" color="#6b7280">
                Documento gerado automaticamente pelo LogiStock.
            </font>
        </para>
        """,
        styles["Normal"]
    )

    elementos.append(rodape)

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = (
        f"inline; filename=baixa_tecnica_{baixa.id}.pdf"
    )

    return response


@bp_baixa_desktop.route("/api/pendentes/count")
def api_baixas_pendentes_count():
    count = (
        BaixaTecnica.query
        .filter(BaixaTecnica.status.in_(["pendente", "pendente_ajuste"]))
        .count()
    )

    return {"count": count}
